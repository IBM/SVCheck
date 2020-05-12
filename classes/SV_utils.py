import requests
import json
import socket
import datetime
import sys
import os
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd

class SV_system(object):
    """
        RESTful client for Spectrum Virtualize API

        __init__(self, IP, username, password, output_dir = "./output/", verbose = False)
        
        It allows to call commands via API run_command("command")

        It allows to save to Excel the current status with generate_excel_report()
    """

    # We need this to silent self signed certs, not good from security point of view
    requests.packages.urllib3.disable_warnings()


    def __init__(self, IP, username, password, output_dir = "./output/", verbose = False):
        self.IP = IP
        self.port = "7443"
        self.base_url = "https://" + self.IP + ":" + self.port + "/rest/"
        self.username = username
        self.password = password
        self.output_dir = output_dir
        self.verbose = verbose
        self.timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        self.master_file = self.output_dir + 'SVCheck_' + self.IP + '_' + self.timestamp + '.xlsx'
        self.log_file = self.output_dir +  'SVCheck_' + self.IP + '_' + self.timestamp + '.log'
        self.SVC_log = self.__set_logger_up()
        self.token_id = self.__get_token()
        self.user_role = self.__get_user_role()
        self.first_run = True
        self._report_commands = [
                                "lssystem",
                                "lsnodecanister",
                                "lssystemstats",
                                "lsnodestats",
                                "lsvdisk", 
                                "lshost",
                                "lshostcluster",
                                "lshostvdiskmap",
                                "lshostclustervolumemap",
                                "lsvdiskaccess",
                                "lsvdiskcopy",
                                "lsportfc",
                                "lsfcconsistgrp",
                                "lsiogrp",
                                "lsmdiskgrp",
                                "lssystemip",
                                "lspartnership",
                                "lseventlog"]
        

    def run_command(self, command):
        """
        Run a command via the API 
        
        It returns a list of dictionaries or a dictionary depending of the command issued
        """
        self.__check_connection()
        has_right = self.__check_user_rights(command)
        command_headers = {'X-Auth-Token': self.token_id}
        command_url = self.base_url + command
        self.SVC_log.debug("Going to send command run " + command + " to API")
        cmd_r = requests.post(command_url, headers=command_headers, verify=False)
        if cmd_r.status_code == 200:
            cmd_d = json.loads(cmd_r.text)
            self.SVC_log.debug("Got HTTP 200 on command run " + command + " to API")
            return cmd_d
        else:
            if has_right: # Might be a valid command and we fail
                self.SVC_log.error("Cannot run " + command + " to API")
                sys.exit(3)
            else:
                self.SVC_log.error("Cannot run " + command + " to API, likely is not a correct command")
                sys.exit(3)


    def generate_excel_report(self):
        """
        Generates the Excel report
        """
        # Generate sheets from list of commands
        for command in self._report_commands:
            self.__parse_command_to_excel(command)
        self.SVC_log.info("Succesfully generated " + self.master_file + " report")


    def __get_user_role(self):
        self.SVC_log.debug("Going to query role of " + self.username)
        current_user = self.run_command("lscurrentuser")
        current_user_role = current_user[1]["role"]
        self.SVC_log.debug("Got role of " + self.username + " is " + current_user_role)
        return current_user_role


    def __check_user_rights(self, command):
        # This is not an exact map to reality as some commands 
        # can be run by specific user that will fail here
        CopyOperator_roles = ["Administrator", "SecurityAdmin", "CopyOperator", "SecurityAdmin"]
        Admin_roles = ["Administrator", "SecurityAdmin", "SecurityAdmin"]
        self.SVC_log.debug("Checking user " + self.username + " has appropiate role for command " + command)
        if command.startswith("ls"):
            self.SVC_log.debug("All groups can run " + command)
            return True
        elif(
            command.startswith("start") or 
            command.startswith("stop") or
            command.startswith("prestart") or
            command.startswith("prestop")
        ):

            if self.user_role in CopyOperator_roles:
                self.SVC_log.debug("Group " + self.user_role + " can run " + command)
                return True
            else:
                self.SVC_log.error("Group " + self.user_role + " cannot run " + command)
                sys.exit(5)
        elif(
            command.startswith("add") or
            command.startswith("ch") or
            command == "expandvdisksize" or
            command.startswith("mk") or
            command == "movevdisk" or
            command.startswith("rm")
        ):
            if self.user_role in Admin_roles:
                self.SVC_log.debug("Group " + self.user_role + " can run " + command)
                return True
            else:
                self.SVC_log.error("Group " + self.user_role + " cannot run " + command)
                sys.exit(5)
        else: # Maybe not a valid command
            self.SVC_log.debug("Cannot match command " + command + " with any know command")
            return False
            

    def __create_output_dir(self):
        # Lets create the output dir
        if os.path.isdir(self.output_dir) == False:
            try:
                os.makedirs(self.output_dir)
                return True
            except BaseException:
                sys.exit(self.__add_timestamp() + ": Cannot create " + self.output_dir)
        

    def __set_logger_up(self):
        self.__create_output_dir()
        sv_log_format = '%(asctime)s %(levelname)-4s:\t %(message)s'
        logging.basicConfig(level=logging.DEBUG,
                    format=sv_log_format,
                    filename=self.log_file,
                    filemode='w')

        console = logging.StreamHandler()
        if self.verbose:
            console.setLevel(logging.DEBUG)
        else:
            console.setLevel(logging.INFO)
        console.setFormatter(logging.Formatter(sv_log_format))
        logging.getLogger('').addHandler(console)
        SVC_log = logging.getLogger(self.IP)
        return SVC_log


    def __add_timestamp(self):
        # We only use this for errors before the log is open
        time_now = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        return self.IP + "_" + time_now


    def __check_connection(self):
        sv_api_check = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sv_api_check.settimeout(2.0)
        location = (self.IP, int(self.port))
        self.SVC_log.debug("Starting check access to API port " + self.IP + ":" + self.port)
        open_port = sv_api_check.connect_ex(location)
        sv_api_check.shutdown(socket.SHUT_RDWR)
        sv_api_check.close()
        
        if open_port != 0:
            self.SVC_log.error("API port " + self.port + " cannot be reached for " + self.IP)    
            sys.exit(6)
        else:
            self.SVC_log.debug("Completed check access to API port "  + self.IP + ":" + self.port)


    def __get_token(self):
        self.__check_connection()
        # We need to encode utf-8 here at least for non ASCII passwords
        token_headers = {'X-Auth-Username': self.username, 'X-Auth-Password': self.password, 'charset': 'utf-8'}
        auth_url = self.base_url + 'auth'
        self.SVC_log.debug("Getting auth token from " + self.IP)
        token_r = requests.post(auth_url, headers=token_headers, verify=False)
        if token_r.status_code == 200:
            token_d = json.loads(token_r.text)
            self.SVC_log.info("Got valid auth token from " + self.IP)
            return token_d['token']
        else:
            self.SVC_log.error("Cannot get auth token with these credentials")
            sys.exit(4)


    def __format_lssystem_to_excel(self):
        
        lssystem = self.run_command("lssystem")
        lssystem_data = []
        lssystem_data.append({
                        "Product name": lssystem['product_name'],
                        "Product model": lssystem['name'],
                        "Serial": lssystem['id'],
                        "Code level": lssystem['code_level'],
                        "Console IP": lssystem['console_IP'],
                        "Contact organization": lssystem['email_organization'],
                        "Contact name": lssystem['email_contact'],
                        "Contact email": lssystem['email_reply'],
                        "Contact phone": lssystem['email_contact_primary'],
                        "Auth service": lssystem['auth_service_configured'],
                        "Auth service type": lssystem['auth_service_type'],
                        "Callhome": lssystem['enhanced_callhome'],
                        "Callhome censor": lssystem['censor_callhome'],
                        # "Quorum lease": lssystem['quorum_lease'],
                        "Copy rate": lssystem['relationship_bandwidth_limit'],
                        "Local raw capacity": lssystem['total_drive_raw_capacity'],
                        "Physical total": lssystem['physical_capacity'],
                        "Physical free": lssystem['physical_free_capacity'],
                        "Easy tier": lssystem['easy_tier_acceleration'],
                        # "NAS key": lssystem['has_nas_key'],
                        "Compression": lssystem['compression_active'],
                        "Compressed virtual": lssystem['compression_virtual_capacity'],
                        "Compressed capacity": lssystem['compression_compressed_capacity'],
                        "Uncompressed capacity": lssystem['compression_uncompressed_capacity'],
                        "Deduplication savings": lssystem['deduplication_capacity_saving'],
                        "Cache prefetch": lssystem['cache_prefetch']
                        })
        # Different models / levels report different tiers
        for tier in lssystem['tiers']:
            tier_key_total = tier['tier']+"_total"
            tier_key_free = tier['tier']+"_free"
            lssystem_data[0].update({tier_key_total: tier['tier_capacity']})
            lssystem_data[0].update({tier_key_free: tier['tier_free_capacity']})
        return lssystem_data


    def __parse_command_to_excel(self, command):

        if self.first_run:
            wb = Workbook()
            ws = wb.active
            ws.title = command
            self.first_run = False
        else:
            wb = load_workbook(self.master_file)
            wb.create_sheet(command)
            ws = wb[command]
        self.SVC_log.debug("Gathering " + command + " information from " + self.IP)
        if command == "lssystem":
            cmd_list = self.__format_lssystem_to_excel()
        else:
            cmd_list = self.run_command(command)
        self.SVC_log.debug("Completed gathering " + command + " information from " + self.IP)

        self.SVC_log.debug("Loading " + command + " information on Pandas dataframe")
        df = pd.DataFrame.from_dict(cmd_list)
        self.SVC_log.debug("Loaded " + command + " information on Pandas dataframe")

        #self.SVC_log.info("Creating " + command + " information into Excel file")

        # Import the data to sheet
        self.SVC_log.debug("Starting loading " + command + " information into Excel file")
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        self.SVC_log.debug("Completed loading " + command + " information into Excel file")

        #self.SVC_log.info("Completed creating " + command + " information into Excel file")
        
        self.__format_sheet(ws, command)

        self.__save_excel(wb, command)


    def __format_sheet(self, ws, command):      
        self.SVC_log.debug("Starting formating " + command + " information into Excel file")

        # Fix the width
        self.SVC_log.debug("Starting fixing width for " + command + " sheet")
        dim_holder = DimensionHolder(worksheet=ws)

        for the_col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(the_col)] = ColumnDimension(ws, min=the_col, max=the_col, width=25)
            ws.column_dimensions = dim_holder
        self.SVC_log.debug("Completed fixing width for " + command + " sheet")

        # Set the header
        self.SVC_log.debug("Starting setting header for " + command + " sheet")
        HEADER_FONT = Font(name='Calibri',
                        size=12,
                        bold=True,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color='000000')
        for the_col in range(ws.min_column, ws.max_column + 1):
            ws.cell(row=1, column=the_col).font = HEADER_FONT
            ws.cell(row=1, column=the_col).fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type = 'darkGrid')
        self.SVC_log.debug("Completed setting header for " + command + " sheet")

        # Center all
        self.SVC_log.debug("Starting centering cells for " + command + " sheet")
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(horizontal='center',
                                            vertical='center')
        self.SVC_log.debug("Completed centering cells for " + command + " sheet")

        # Create alternate colors
        self.SVC_log.debug("Starting coloring alternate rows for " + command + " sheet")
        for the_row in range(2, ws.max_row + 1, 2):
            for the_col in range(1, ws.max_column + 1):
                ws.cell(row=the_row, column=the_col).fill = PatternFill(start_color='66cc00', end_color='66cc00', fill_type = 'darkGrid')

        for the_row in range(3, ws.max_row + 1, 2):
            for the_col in range(1, ws.max_column + 1):
                ws.cell(row=the_row, column=the_col).fill = PatternFill(start_color='b3ff66', end_color='b3ff66', fill_type = 'lightGrid')
        self.SVC_log.debug("Completed coloring alternate rows for " + command + " sheet")

        self.SVC_log.debug("Completed formating " + command + " information into Excel file")


    def __save_excel(self, wb, command):
        # Save and close
        self.SVC_log.debug("Starting saving " + command + " information into Excel file")
        try:
            wb.save(self.master_file)
        except BaseException:
            self.SVC_log.error("Cannot write " + self.master_file)
            sys.exit(2)
        self.SVC_log.info("Completed saving " + command + " information into Excel file")
